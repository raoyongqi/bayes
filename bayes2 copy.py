import pandas as pd
import numpy as np
from scipy.stats import spearmanr
from openpyxl import Workbook
from openpyxl.styles import Font

# 文件路径
file_path = "C:/Users/r/Desktop/bayes/data/selection.csv"  # 替换为你的文件路径

# 读取数据
selection = pd.read_csv(file_path)

# 删除 'RATIO' 列
selection['PL'] = selection['RATIO']
selection = selection.drop(columns=['RATIO'])
selection.columns = selection.columns.map(lambda x: x.replace('_', r' '))

# 计算 Spearman 相关系数矩阵和 p 值矩阵
corr_matrix = selection.corr(method='spearman')  # Spearman 相关系数矩阵
p_matrix = selection.corr(method=lambda x, y: spearmanr(x, y)[1])  # p 值矩阵

# 格式化相关系数矩阵
formatted_matrix = corr_matrix.copy()
for i in range(len(corr_matrix)):
    for j in range(len(corr_matrix.columns)):
        # 保留两位小数并添加显著性标记
        value = f"{corr_matrix.iloc[i, j]:.2f}"
        
        # 添加显著性标记
        if p_matrix.iloc[i, j] < 0.05:
            value += "*"
        
        formatted_matrix.iloc[i, j] = value

# 创建掩码，只显示下三角部分
mask = np.triu(np.ones_like(corr_matrix, dtype=bool), k=0)  # k=0 包括对角线

# 提取下三角部分的相关系数（不包括对角线）
lower_triangle_corr = formatted_matrix.where(~mask)

heatmap_data = lower_triangle_corr.dropna(how='all', axis=0).dropna(how='all', axis=1)
formatted_df = heatmap_data.fillna("").astype(str)

# 创建一个新的 Excel 文件并使用 openpyxl 写入格式化数据
output_file = "C:/Users/r/Desktop/bayes/formatted_correlation_matrix.xlsx"  # 选择输出文件路径
wb = Workbook()
ws = wb.active

# 写入列头
for col_idx, col_name in enumerate(formatted_df.columns, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

# 写入数据并进行加粗
for row_idx, row in enumerate(formatted_df.itertuples(index=False), 2):
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        value_no_star = value.replace('*', '')  # 移除 '*'

        print(value_no_star)
  
        try:
            # 确保 value_no_star 不是空字符串，再进行加粗判断
            if value_no_star and float(value_no_star) < 0:
                cell.font = Font(bold=True)
        except ValueError:
            # 如果转换失败（可能是空值或其他非数字的情况），跳过
            pass

# 保存文件
wb.save(output_file)
